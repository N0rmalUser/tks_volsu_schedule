# TKS VOLSU SCHEDULE BOT
# Copyright (C) 2024 N0rmalUser
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program. If not, see <http://www.gnu.org/licenses/>.

import logging
import sqlite3

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from pandas import concat, read_sql_query

from app.config import ROOMS_SHEETS_PATH, SCHEDULE_DB, TEACHERS_SHEETS_PATH, STUDENTS


def teacher(teacher_name: str) -> str:
    days = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота"]
    times = ["08", "10", "12", "13", "15", "17", "18"]

    start_row = 2
    start_col = 3
    conn = sqlite3.connect(SCHEDULE_DB)

    query = """
        SELECT
            s.Time,
            s.DayOfWeek,
            s.WeekType,
            GROUP_CONCAT(g.GroupName, ', ') AS GroupNames,
            r.RoomName,
            sub.SubjectName,
            Subgroup
        FROM (
            SELECT Time, DayOfWeek, WeekType, SubjectID, GroupID, RoomID, TeacherID, Subgroup FROM Schedule
            UNION ALL
            SELECT Time, DayOfWeek, WeekType, SubjectID, GroupID, RoomID, TeacherID, 0 AS Subgroup FROM CollegeSchedule
        ) s
        JOIN Subjects sub ON s.SubjectID = sub.SubjectID
        JOIN Groups g ON s.GroupID = g.GroupID
        JOIN Rooms r ON s.RoomID = r.RoomID
        JOIN Teachers t ON s.TeacherID = t.TeacherID
        WHERE t.TeacherName = ?
        GROUP BY s.Time, s.DayOfWeek, s.WeekType, r.RoomName, sub.SubjectName;
        """

    teacher_df = read_sql_query(sql=query, con=conn, params=[teacher_name])
    student_df = pd.DataFrame()
    if teacher_name in STUDENTS.keys():
        query = """
                SELECT
                    s.Time,
                    s.DayOfWeek,
                    s.WeekType,
                    t.TeacherName AS GroupNames,
                    r.RoomName,
                    sub.SubjectName,
                    0 AS Subgroup
                FROM (
                    SELECT Time, DayOfWeek, WeekType, SubjectID, GroupID, RoomID, TeacherID FROM Schedule
                    UNION ALL
                    SELECT Time, DayOfWeek, WeekType, SubjectID, GroupID, RoomID, TeacherID FROM CollegeSchedule
                ) s
                JOIN Subjects sub ON s.SubjectID = sub.SubjectID
                JOIN Groups g ON s.GroupID = g.GroupID
                JOIN Rooms r ON s.RoomID = r.RoomID
                JOIN Teachers t ON s.TeacherID = t.TeacherID
                WHERE g.GroupName = ?
                ORDER BY s.Time, s.DayOfWeek, s.WeekType, r.RoomName, sub.SubjectName;
                """
        student_df = read_sql_query(sql=query, con=conn, params=[STUDENTS[teacher_name]])

    conn.close()
    df = pd.concat([teacher_df, student_df], axis=0)
    wb = load_workbook(TEACHERS_SHEETS_PATH / "template.xlsx")
    ws = wb.active
    not_empty_rows = []
    for _, data in df.iterrows():
        week_type = data["WeekType"]
        time = times.index(data["Time"][:2])
        row = start_row + days.index(data["DayOfWeek"]) * 14 + time * 2

        if week_type != "Числитель":
            row += 1

        not_empty_rows.append(row)
        if week_type == "Числитель":
            not_empty_rows.append(row + 1)
        else:
            not_empty_rows.append(row - 1)

        ws.cell(
            row=row, column=start_col
        ).value = f"{data['GroupNames']}{f'.{data["Subgroup"]}' if data['Subgroup'] else ''}"
        ws.cell(row=row, column=start_col + 1).value = data["RoomName"]
        ws.cell(row=row, column=start_col + 2).value = data["SubjectName"]

    for row in range(ws.max_row, start_row - 1, -1):
        if row not in not_empty_rows:
            ws.delete_rows(row)

    for row in range(start_row, ws.max_row, 2):
        for col in range(2, start_col + 3):
            if ws.cell(row=row, column=col).value == ws.cell(row=row + 1, column=col).value:
                ws.merge_cells(start_row=row, start_column=col, end_row=row + 1, end_column=col)

    prev_value = None
    start_merge_row = start_row

    for row in range(start_row, ws.max_row + 2):
        cell = ws.cell(row=row, column=1)
        if cell.value != prev_value and prev_value is not None:
            ws.merge_cells(start_row=start_merge_row, start_column=1, end_row=row - 1, end_column=1)
            start_merge_row = row
        prev_value = cell.value

    for row in range(start_row, ws.max_row):
        for col in range(1, start_col + 3):
            if row % 2 == 0:
                ws.cell(row=row, column=col).border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    bottom=Side(style="dotted"),
                )
            else:
                ws.cell(row=row, column=col).border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    bottom=Side(style="thin"),
                )
        if ws.cell(row=row, column=1).value is not None:
            for col in range(1, start_col + 3):
                ws.cell(row=row, column=col).border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="double"),
                    bottom=Side(style="dotted"),
                )
    file_name = TEACHERS_SHEETS_PATH / f"{teacher_name}.xlsx"
    wb.save(file_name)
    logging.info(f"Создан файл расписания {file_name}")
    return str(file_name)


def room(room_name: str):
    days = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота"]
    times = ["08", "10", "12", "13", "15", "17", "18"]

    start_row = 2
    start_col = 3

    schedule_dfs = []

    conn = sqlite3.connect(SCHEDULE_DB)
    query = """
        SELECT
            s.Time,
            s.DayOfWeek,
            s.WeekType,
            GROUP_CONCAT(g.GroupName, ', ') AS GroupNames,
            t.TeacherName,
            sub.SubjectName,
            r.RoomName
        FROM (
            SELECT Time, DayOfWeek, WeekType, SubjectID, GroupID, RoomID, TeacherID FROM Schedule
            UNION ALL
            SELECT Time, DayOfWeek, WeekType, SubjectID, GroupID, RoomID, TeacherID FROM CollegeSchedule
        ) s
        JOIN Subjects sub ON s.SubjectID = sub.SubjectID
        JOIN Groups g ON s.GroupID = g.GroupID
        JOIN Rooms r ON s.RoomID = r.RoomID
        JOIN Teachers t ON s.TeacherID = t.TeacherID
        WHERE r.RoomName = ?
        GROUP BY s.Time, s.DayOfWeek, s.WeekType, t.TeacherName, sub.SubjectName, r.RoomName;
    """

    if room_name[:-1] == "2-13":
        room_variants = [room_name[:-1] + variant for variant in ["М", "аМ", "бМ"]]
    elif room_name[:-1] == "3-15":
        room_variants = [room_name[:-1] + variant for variant in ["К", "аК", "бК"]]
    else:
        room_variants = [room_name]

    for variant in room_variants:
        df = read_sql_query(sql=query, con=conn, params=[variant])
        schedule_dfs.append(df)

    df = concat(schedule_dfs, ignore_index=True)
    conn.close()

    wb = load_workbook(ROOMS_SHEETS_PATH / "template.xlsx")
    ws = wb.active
    not_empty_rows = []

    for _, data in df.iterrows():
        week_type = data["WeekType"]
        time = times.index(data["Time"][:2])
        row = start_row + days.index(data["DayOfWeek"]) * 14 + time * 2

        if week_type != "Числитель":
            row += 1

        not_empty_rows.append(row)
        if week_type == "Числитель":
            not_empty_rows.append(row + 1)
        else:
            not_empty_rows.append(row - 1)

        room_index = data["RoomName"][-2]
        if room_index.isdigit():
            room_index = ""
        else:
            room_index = f" ({room_index})"

        teacher_cell = ws.cell(row=row, column=start_col)
        teacher_name_with_room = f"{data['TeacherName']}{room_index}"

        if teacher_cell.value:
            teacher_cell.value += f",\n{teacher_name_with_room}"
        else:
            teacher_cell.value = teacher_name_with_room

        subject_cell = ws.cell(row=row, column=start_col + 1)
        if subject_cell.value:
            subject_cell.value += f",\n{data['SubjectName']}"
        else:
            subject_cell.value = data["SubjectName"]

    for row in range(ws.max_row, start_row - 1, -1):
        if row not in not_empty_rows:
            ws.delete_rows(row)

    for row in range(start_row, ws.max_row, 2):
        for col in range(2, start_col + 2):
            if ws.cell(row=row, column=col).value == ws.cell(row=row + 1, column=col).value:
                ws.merge_cells(start_row=row, start_column=col, end_row=row + 1, end_column=col)

    prev_value = None
    start_merge_row = start_row

    for row in range(start_row, ws.max_row + 2):
        cell = ws.cell(row=row, column=1)
        if cell.value != prev_value and prev_value is not None:
            ws.merge_cells(start_row=start_merge_row, start_column=1, end_row=row - 1, end_column=1)
            start_merge_row = row
        prev_value = cell.value

    for row in range(start_row, ws.max_row):
        for col in range(1, start_col + 2):
            if row % 2 == 0:
                ws.cell(row=row, column=col).border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    bottom=Side(style="dotted"),
                )
            else:
                ws.cell(row=row, column=col).border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    bottom=Side(style="thin"),
                )
        if ws.cell(row=row, column=1).value is not None:
            for col in range(1, start_col + 2):
                ws.cell(row=row, column=col).border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="double"),
                    bottom=Side(style="dotted"),
                )
    file_name = ROOMS_SHEETS_PATH / f"{room_name}.xlsx"
    wb.save(file_name)
    logging.info(f"Создан файл расписания {file_name}")
    return str(file_name)
